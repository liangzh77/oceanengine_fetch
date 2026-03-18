"""数据提取模块：组织切换、通过下载 Excel 获取表格数据"""
import logging
import os
import tempfile

import openpyxl
from playwright.sync_api import Page

logger = logging.getLogger(__name__)


class DataExtractor:
    def __init__(self, page: Page):
        self.page = page

    def switch_organization(self, org_name: str, app_name: str):
        """切换组织：打开组织面板，搜索app名称，点击对应节点"""
        logger.info(f"switch org: {org_name} - {app_name}")

        org_btn = self.page.locator(".ebp-header-account-info").first
        org_btn.click()
        self.page.wait_for_timeout(2000)

        search_input = self.page.locator(
            ".ebp-team-switch-opper-container input[placeholder*='组织名称']"
        ).first
        search_input.click()
        search_input.fill(app_name)
        self.page.wait_for_timeout(2000)

        cards = self.page.locator(".node-card .node-card-part-content-name").all()
        clicked = False
        for card in cards:
            if card.is_visible() and app_name in card.inner_text():
                card.click()
                clicked = True
                break

        if not clicked:
            self.page.get_by_text(app_name, exact=False).first.click()

        self.page.wait_for_timeout(3000)
        try:
            self.page.wait_for_selector("table.ovui-table", timeout=30000)
            self.page.wait_for_timeout(3000)
        except Exception:
            logger.warning("table not found after org switch, waiting more...")
            self.page.wait_for_timeout(10000)
        logger.info(f"switched to: {org_name} - {app_name}")

    def set_date_filter(self, target_date: str):
        """设置日期筛选器，起止日期均设为 target_date（YYYY-MM-DD）"""
        logger.info(f"设置日期筛选器: {target_date}")

        # 点击开始日期输入框，打开日历
        self.page.locator("input[placeholder*='开始日期']").first.click()
        self.page.wait_for_timeout(800)

        # 选开始日期
        self._pick_date(target_date)
        self.page.wait_for_timeout(400)

        # 日历自动切换到结束日期，选同一天
        self._pick_date(target_date)
        self.page.wait_for_timeout(800)

        logger.info(f"日期筛选器已设置: {target_date}")

    def _pick_date(self, target_date: str):
        """在已打开的日历中点击指定日期（title=YYYY-MM-DD），必要时翻月"""
        popper_sel = ".ovui-range-picker__popper--show"
        prev_btn_sel = ".ovui-date__header-prev-month"
        next_btn_sel = ".ovui-date__header-next-month"

        for _ in range(24):
            # 先尝试直接点击目标日期单元格
            cell = self.page.locator(f'{popper_sel} td[title="{target_date}"]').first
            if cell.is_visible():
                cell.click()
                return

            # 判断需要往前还是往后翻月
            # 读取当前日历第一个面板显示的年月
            cur_ym = self.page.evaluate(f"""
                () => {{
                    const popper = document.querySelector('{popper_sel}');
                    if (!popper) return null;
                    const header = popper.querySelector('.ovui-date__header');
                    if (!header) return null;
                    const spans = header.querySelectorAll('span');
                    for (const s of spans) {{
                        const m = s.textContent.match(/(\\d{{4}}).*?(\\d{{1,2}})/);
                        if (m) return [parseInt(m[1]), parseInt(m[2])];
                    }}
                    return null;
                }}
            """)
            if not cur_ym:
                logger.warning("无法读取日历年月，跳过翻月")
                break

            from datetime import datetime
            target_dt = datetime.strptime(target_date, "%Y-%m-%d")
            cur_year, cur_month = cur_ym
            if (cur_year, cur_month) > (target_dt.year, target_dt.month):
                self.page.locator(f"{popper_sel} {prev_btn_sel}").first.click()
            else:
                self.page.locator(f"{popper_sel} {next_btn_sel}").last.click()
            self.page.wait_for_timeout(300)

        logger.warning(f"未能在日历中找到日期: {target_date}")

    def _click_tab(self, tab_name: str):
        """点击指定的 tab 页（账户/项目/单元）"""
        logger.info(f"click tab: {tab_name}")
        tab = self.page.locator(f"text='{tab_name}'").first
        tab.click()
        self.page.wait_for_timeout(3000)
        try:
            self.page.wait_for_selector("table.ovui-table", timeout=15000)
            self.page.wait_for_timeout(2000)
        except Exception:
            logger.warning(f"tab '{tab_name}' table not found after click")

    def _download_excel(self) -> list[dict]:
        """点击下载按钮，用 Playwright download API 获取 Excel"""
        download_btn = self.page.locator("iconpark-icon[name='oc-icon-download']").first
        if not download_btn.is_visible():
            logger.warning("download button not found")
            return []

        # 用 Playwright 的 expect_download 捕获下载（数据量大时增加超时）
        with self.page.expect_download(timeout=120000) as download_info:
            download_btn.locator("..").click()

        download = download_info.value
        # 保存到临时文件
        tmp_path = os.path.join(tempfile.gettempdir(), download.suggested_filename)
        download.save_as(tmp_path)
        logger.info(f"downloaded: {download.suggested_filename} ({os.path.getsize(tmp_path)} bytes)")

        rows = self._read_excel(tmp_path)

        try:
            os.remove(tmp_path)
        except Exception:
            pass

        return rows

    @staticmethod
    def _read_excel(path: str) -> list[dict]:
        """读取 Excel 文件，返回 list[dict]，键为列名"""
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            record = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    record[headers[i]] = value
            if any(v for v in record.values()):
                rows.append(record)
        return rows

    def fetch_accounts(self) -> list[dict]:
        self._click_tab("账户")
        return self._download_excel()

    def fetch_projects(self) -> list[dict]:
        self._click_tab("项目")
        return self._download_excel()

    def fetch_units(self) -> list[dict]:
        self._click_tab("单元")
        rows = self._download_excel()
        if rows:
            # 临时日志：打印单元的列名和前几行的状态值
            logger.info(f"单元Excel列名: {list(rows[0].keys())}")
            status_vals = set()
            for r in rows:
                for k, v in r.items():
                    if '状态' in str(k):
                        status_vals.add(f"{k}={v}")
            logger.info(f"单元状态值样本: {list(status_vals)[:10]}")
        return rows

    def fetch_all(self) -> dict:
        return {
            "accounts": self.fetch_accounts(),
            "projects": self.fetch_projects(),
            "units": self.fetch_units(),
        }
